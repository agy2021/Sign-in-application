from openpyxl import *
from tkinter import *
import time

def main():
    wb_name = "Users.xlsx"
    wb = load_workbook(wb_name)
    ws = wb.active

    win = Tk()
    win.title("Sign up or sign in? ")
    win.geometry("500x250")

    def up_or_in():
        a = winentry.get()
        a.lower()

        if a == "sign up":
            def save():
                username_val = username.get()
                password_val = password.get()

                ws.append([username_val, password_val])
                wb.save(wb_name)

                w = Tk()
                w.title("Saved!")
                w.geometry("500x250")
                l1 = Label(w, text="Username and password saved! Restarting program...")
                l1.pack()
                w.mainloop()
                time.sleep(1.5)
                w.destroy()
                main()

            win1 = Tk()    
            win1.title("Enter new username and password.")
            win1.geometry("500x250")

            win1label = Label(win1, text="Enter your new username:")
            win1label.pack()
            username = Entry(win1)
            username.pack()

            win1label = Label(win1, text="Enter your new password:")
            win1label.pack()
            password = Entry(win1)
            password.pack()

            win1btn = Button(win1, text="Save", command=lambda:save())
            win1btn.pack()

        elif a == "sign in":
            win2 = Tk()
            win2.title("Enter username and password.")
            win2.geometry("500x250")
            username = Entry(win2)
            password = Entry(win2)
            
            def check():
                username_val = username.get()
                password_val = password.get()
                correct2 = False
                correct1 = False

                for a in ws["A"]:
                    if a.value in username_val:
                        correct1 = True
                for a in ws["B"]:
                    if a.value in password_val:
                        correct2 = True

                if correct1 and correct2:
                        correct = Tk()
                        correct.title("Correct username and password!")
                        correct.geometry("500x250")

                        correct_text = Label(correct, text="Welcome!")
                        correct_text.pack()

                        correct.mainloop()
                else:
                        incorrect = Tk()
                        incorrect.title("Incorrect.")
                        incorrect.geometry("500x250")

                        incorrect_text = Label(incorrect, text="Incorrect credentials.")
                        incorrect_text.pack()

                        incorrect.mainloop()

            win2label = Label(win2, text="Enter username:")
            win2label.pack()
            username.pack()
            win2label2 = Label(win2, text="Enter password:")
            win2label2.pack()
            password.pack()

            win2btn = Button(win2, text="Go", command=lambda:check())
            win2btn.pack()

            wb.save(wb_name)
            win2.mainloop()

    winlabel = Label(win, text="Would you like to sign up or sign in? ")
    winlabel.pack()

    winentry = Entry(win)
    winentry.pack()

    winbtn = Button(win, text="Go", command=lambda:up_or_in())
    winbtn.pack()

    wb.save(wb_name)
    win.mainloop()